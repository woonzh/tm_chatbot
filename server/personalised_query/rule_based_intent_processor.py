from openpyxl import load_workbook
import argparse

'''
wifi column ids:
----------------
1 -> Chatter Name
2 -> Service Number 
3 -> Account Name (Owner)
4 -> BRN Number
5 -> Package Subscribe
6 -> Service Activation
7 -> Order Number
8 -> Service Application
9 -> Contract End Period
10 -> Mobile Number
11 -> Account Number
12 -> Billing Address
13 -> Installation Address
14 -> E-mail Address
15 -> Bill Amount
16 -> Bill Date
17 -> Bill Due Date
18 -> Outstanding Payment
19 -> Credit Limit
20 -> Turbo Upgraded
21 -> Trouble Ticket No
22 -> Trouble Ticket Created Date
23 -> Restoration Appointment Date
24 -> Technical Name (Restorer)
25 -> Technical Contact Details

mobile column ids:
------------------
1 -> Full Name :
2 -> ID Type :
3 -> ID Type :
4 -> Mobile Phone :
5 -> Nationality :
6 -> Email :
7 -> Address Line 1:
8 -> Address Line 2:
9 -> City:
10 -> Post Code:
11 -> State:
12 -> Account Code :
13 -> Outstanding / Balance :
14 -> Status:
15 -> Registration Date:
16 -> Billing Cycle:
17 -> Activation Date:
18 -> Account Flag:
19 -> Account Type:
20 -> Contact Number:
21 -> Auto Pay:
22 -> Used Credit Limit :
23 -> Default Credit Limit :
24 -> APN :
25 -> subscription Line :
26 -> Day 1 Active Date :
27 -> last Prepaid Pass Purchase date :
28 -> Last Transaction Date:
29 -> Last Reload Date :
30 -> PUK Number:
31 -> APN :
32 -> Roaming :
33 -> MNP Port In status :
34 -> MNP Port In rejected reason : 

'''

wifi_memory = None
mobile_memory = None

wifi_data_dict_names = [
	'chatter_name', 
	'service_num', 
	'account_name_owner', 
	'brn_num', 
	'package_subscribe', 
	'service_activation', 
	'order_num', 
	'service_application', 
	'contract_end_period', 
	'mobile_num', 
	'account_num',
	'billing_addr', 
	'installation_addr', 
	'email_addr', 
	'bill_amt',
	'bill_date',
	'bill_due_date',
	'outstanding_payment',
	'credit_limit',
	'turbo_upgraded',
	'trouble_ticket_no',
	'trouble_ticket_created_date',
	'restoration_appt_date',
	'technical_name_restorer',
	'technical_contact_details'
]

mobile_data_dict_names = [
	'full_name', 
	'id_type',
	'id_num',
	'mobile_num',
	'nationality',
	'email_addr',
	'addr_line_1',
	'addr_line_2',
	'city',
	'postal_code',
	'state',
	'account_code',
	'outstanding_bal',
	'status',
	'registration_date',
	'billing_cycle',
	'activation_date',
	'account_flag',
	'account_type',
	'contact_number',
	'auto_pay',
	'used_credit_limit',
	'default_credit_limit',
	'apn',
	'subscription_line',
	'day_1_active_date',
	'last_prepaid_purchase_date',
	'last_transaction_date',
	'last_reload_date',
	'puk_num',
	'apn',
	'roaming',
	'mnp_port_in_status',
	'mnp_port_in_rejected_reason'
]

def get_data(name):
	# Byan: we want to minimize reads from the excel file, so we grab all data from both wifi and mobile (if they match) and store it in memory
    wb = load_workbook(filename='database/database.xlsx', read_only=True)
	
    wifi_sheet = wb['Wifi Data']
    wifi_data = None
    for row in range(1, wifi_sheet.max_row + 1):
        try:
            if wifi_sheet.cell(row, 1).value.lower() == name.lower():
                wifi_data = {}
                for index in range(1, len(wifi_data_dict_names)+1):
				#print('DEBUG:', wifi_data_dict_names[index], ':', wifi_sheet.cell(row, index).value)
                    wifi_data[wifi_data_dict_names[index-1]] = str(wifi_sheet.cell(row, index).value)
        except:
            pass

    mobile_sheet = wb['Mobile Data']
    mobile_data = None
    for row in range(1, mobile_sheet.max_row + 1):
        try:
            if mobile_sheet.cell(row, 1).value.lower() == name.lower():
                mobile_data = {}
                for index in range(1, len(mobile_data_dict_names)+1):
                    mobile_data[mobile_data_dict_names[index-1]] = str(mobile_sheet.cell(row, index).value)
        except:
            pass

    wb._archive.close()

    return wifi_data, mobile_data

def handle_intent(name, intent):
	global wifi_memory, mobile_memory
	print('INFO: Chatter Name:', name)
	print('INFO: Intent ID:', intent)
#	wifi_data, mobile_data = get_data(name)
#	print('wifi data ', wifi_data)
#	print(' mobile data ', mobile_data)
#	wifi_memory=wifi_data
#	mobile_memory=mobile_data
    
	if intent == 0:
		result = ''
	elif intent in range(1, 9):
		if wifi_memory is None:
			print('INFO:', name, 'is not found in wifi data. Skipping.')
			result='Dear ' + name + ', your name is not found in the wifi customer list.'
		else:
			result = 'Dear ' + name + ', '
			if intent == 1:
				# email, bill amount, bill date, bill due date, service activation
				email_addr = wifi_memory['email_addr']
				bill_amt = wifi_memory['bill_amt']
				bill_date = wifi_memory['bill_date']
				bill_due_date = wifi_memory['bill_due_date']
				service_activation = wifi_memory['service_activation']

				result = result + 'your bill of ' + bill_amt + ' (due ' + bill_due_date + ') will be mailed to ' + email_addr + '.'
			elif intent == 2:
				# bill amount, bill due date, outstanding payment
				bill_amt = wifi_memory['bill_amt']
				bill_due_date = wifi_memory['bill_due_date']
				outstanding_payment = wifi_memory['outstanding_payment']

				result = result + 'you have a bill of ' + bill_amt + ' due on ' + bill_due_date + ' and an outstanding payment of ' + outstanding_payment + '.'
			elif intent == 3:
				# turbo upgraded, contract end period
				turbo_upgraded = wifi_memory['turbo_upgraded']
				contract_end_period = wifi_memory['contract_end_period']
				if turbo_upgraded == 'Not entitle':
					result = result + 'you are not entitled to turbo service upgrade.'
				else:
					result = result + 'you are entitled to turbo service upgrade. Your contract expires on ' + contract_end_period + '.'
				
			elif intent == 4: 
				# trouble ticket no, trouble ticket created date, restoration appointment date, technical name (restorer), technical contact details
				trouble_ticket_no = wifi_memory['trouble_ticket_no']
				trouble_ticket_created_date = wifi_memory['trouble_ticket_created_date']
				restoration_appt_date = wifi_memory['restoration_appt_date']
				technical_name_restorer = wifi_memory['technical_name_restorer']
				technical_contact_details = wifi_memory['technical_contact_details']

				result = 'we have created trouble ticket ' + trouble_ticket_no + ' on ' + trouble_ticket_created_date + ' and will also be sending our technician ' + technical_name_restorer + ' (contact number: ' + technical_contact_details + ') down on ' + restoration_appt_date + ' to inspect your internet service.'

			elif intent == 5:
				# mobile number, restoration appointment date, technical name (restorer), technical contact details
				mobile_num = wifi_memory['mobile_num']
				restoration_appt_date = wifi_memory['restoration_appt_date']
				technical_name_restorer = wifi_memory['technical_name_restorer']
				technical_contact_details = wifi_memory['technical_contact_details']

				result = 'we apologize for not getting back to you sooner. Our customer service officer will contact you at your number ' + mobile_num + ' soon and we will also be sending our technician ' + technical_name_restorer + ' (contact number: ' + technical_contact_details + ') down on ' + restoration_appt_date + ' to inspect your internet service.'
			elif intent == 6:
				# technical name (restorer), technical contact details
				technical_name_restorer = wifi_memory['technical_name_restorer']
				technical_contact_details = wifi_memory['technical_contact_details']

				result = 'your technical name is ' + technical_name_restorer + ' (contact number:' + technical_contact_details + ').'
			elif intent == 7:
				# NA
				result = result + 'please hold while I redirect you to our customer service officer.'
			elif intent == 8: 
				# service activation, contract end period
				service_activation = wifi_memory['service_activation']
				contract_end_period = wifi_memory['contract_end_period']

				result = result + 'your service was activated on ' + service_activation + ' and your contract expires on ' + contract_end_period + '.'
	elif intent in range(9,17):
		if mobile_memory is None:
			print('INFO:', name, 'is not found in mobile data. Skipping.')
			result = result='Dear ' + name + ', your name is not found in the mobile customer list.'
		else:
			result = 'Dear ' + name + ', '
			if intent == 9:
				# outstanding / balance
				outstanding_bal = mobile_memory['outstanding_bal']

				result = result + 'your current outstanding amount is ' + outstanding_bal + '.'
			elif intent == 10:
				# status
				status = mobile_memory['status']

				result = result + 'your service status is currently ' + status + '.'

				if status == 'Active':
					result = result + ' We recommend moving to a different location and seeing if the issue persists. If it does, please contact us again for help.'

			elif intent == 11:
				# auto pay, used credit limit, default credit limit
				auto_pay = mobile_memory['auto_pay']
				used_credit_limit = mobile_memory['used_credit_limit']
				default_credit_limit = mobile_memory['default_credit_limit']

				result = result + 'auto pay service for your account is currently ' + auto_pay + '.'

				if auto_pay == 'Activated':
					result = result + ' You have used ' + used_credit_limit + '/' + default_credit_limit + ' of your credit limit.'

					if int(used_credit_limit) - int(default_credit_limit) > -50:
						result = result + ' Please consider topping up.'
			elif intent == 12:
				# status
				status = mobile_memory['status']

				result = result + 'your service status is currently ' + status + '.'

				if status == 'Active':
					result = result + ' We recommend moving to a different location and seeing if the issue persists. If it does, please contact us again for help.'
			elif intent == 13:
				# last reload date, status
				last_reload_date = mobile_memory['last_reload_date']
				status = mobile_memory['status']

				result = result + 'your last reload date is noted to be on ' + last_reload_date + ' by us. Your service status is currently ' + status + '.'

				if status == 'Active':
					result = result + ' We recommend moving to a different location and seeing if the issue persists. If it does, please contact us again for help.'
			elif intent == 14:
				# status, roaming
				status = mobile_memory['status']
				roaming = mobile_memory['roaming']

				result = result + 'your service status is currently ' + status + ' and your roaming service is ' + roaming + '.'

				if roaming == 'Active':
					result = result + ' We recommend moving to a different location and seeing if the issue persists. If it does, please contact us again for help.'
			elif intent == 15:
				# account code
				account_code = mobile_memory['account_code']

				result = result + 'your unifi account number is ' + account_code + '.'
			elif intent == 16:
				# puk number
				puk_num = mobile_memory['puk_num']

				result = result + 'the PUK code for unblocking your phone is ' + puk_num + '.'

	return result

def init_name(name):
    global wifi_memory, mobile_memory
    wifi_memory, mobile_memory = get_data(name)
    
    exists={}
    if wifi_memory is None:
        exists['wifi']=False
    else:
        exists['wifi']=True
    
    if mobile_memory is None:
        exists['mobile']=False
    else:
        exists['mobile']=True
        
    return exists

def personalized_query(name, intentLists):
	store=[]
	for i in intentLists:
		store.append(handle_intent(name, i))
	# returns a string
	return store

def test(name, idList):
    init_name(name)
    for i in idList:
        try: 
            personalized_query('rec', [i])
        except Exception as e:
            print(i,' error: ', e)
            
#test('mr shah', list(range(1, 17)))

#init_name('mr shah')

#res=personalized_query('rec', [1,2,3])
#print(res)

#if __name__ == "__main__":
#	parser = argparse.ArgumentParser(description='Process docx files')
#	parser.add_argument("-t", dest='testmode', action='store', default=False, help='test mode')
#	parser.add_argument("-n", dest='name', action='store', default='', help='chatter name')
#	parser.add_argument("-i", dest='intent', action='store', default=0, help='intent ID')
#	args = parser.parse_args()
#
#	testmode = args.testmode
#	
#	if testmode:
#		name = 'Mr Shah'
#		
#		wifi_memory, mobile_memory = get_data(name)
#		for intent in range(0, 16):
#			print('OUTPUT:', handle_intent(name, intent))
#	else:
#		name = args.name
#		intent = args.intent